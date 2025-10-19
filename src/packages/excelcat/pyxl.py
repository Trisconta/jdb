# pyxl.py  (c)2025  Henrique Moreira

""" pyxl.py - Simple Excel operations using openpyxl
"""

# pylint: disable=missing-function-docstring

import copy
from json import dumps
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.header_footer import HeaderFooter

VALID_CODINGS = (
    "ascii", "utf-8", "ISO-8859-1",
)


def tester(infile, do_save=True):
    mbk = MyBook()
    isok = mbk.load(infile)
    assert isok, infile
    #sht = mbk.sheet()
    wbk = openpyxl.Workbook()
    new_sheet = wbk.active
    #dct = copycat(sht, new_sheet)
    dct = mbk.copy_sheet(1, new_sheet)
    show_w(dct, "widths")
    if not do_save:
        return False
    mbk.update_scale(new_sheet)
    wbk.save("/tmp/new.xlsx")
    return True

def show_w(dct, what):
    for key, item in dct[what].items():
        print(key, item)


class GenData:
    """ Abstract class for Generic Data
    """
    def __init__(self, data, name, encoding=None):
        self.name = name
        self._data, self._encoding = data, "ascii"
        self._enc_type = self.set_encoding(encoding)

    def raw(self):
        return self._data

    def set_encoding(self, encoding) -> int:
        """ Checks whether string encoding is valid. """
        if encoding is None:
            return "ascii"
        try:
            idx = VALID_CODINGS.index(encoding)
        except ValueError:
            idx = -1
        return idx

    def conv_to_pixels(self, x_width, factor=7):
        """ Converts Excel widths stored into pixels, approximately. """
        return conv_x_to_pixels(x_width, factor)

    def __str__(self) -> str:
        return self.to_string()

    def to_string(self) -> str:
        return self._dump_json_string()

    def _dump_json_string(self, ensure_ascii=True):
        """ Dump JSON from data
        """
        cont = self._data	# Content!
        ind, asort, ensure = 2, True, ensure_ascii
        astr = dumps(cont, indent=ind, sort_keys=asort, ensure_ascii=ensure)
        return astr


class MyBook(GenData):
    """ Excel book manipulation.
    """
    def __init__(self, data=None, name="B", encoding=None):
        super().__init__(
            {} if data is None else data, name, encoding,
        )
        self._read_only = False
        self._book = None
        self._info = {}

    def book(self):
        assert self._book is not None, "book?"
        return self._book

    def sheet(self, which=0):
        """ Returns the sheet by 1..n index; if 0 chosen, returns the first one.
        """
        idx = which - 1 if which > 0 else 0
        name = self._book.sheetnames[idx]
        return self._book[name]

    def load(self, path:str) -> bool:
        self._data = {}
        try:
            wb1 = openpyxl.open(path, read_only=self._read_only, data_only=True)
        except FileNotFoundError:
            return False
        self._book = wb1
        return True

    def copy_sheet(self, index:int, new_sheet):
        assert 0 <= index <= len(self._book.worksheets), "copy_sheet() out of range"
        sht = self.sheet(index)
        dct = copycat(sht, new_sheet)
        self._info = dct
        return dct

    def update_scale(self, sheet):
        """ Automatically updates the scale of 'sheet' (as output) for printing (A4). """
        w_vals = self._info["widths"]
        widths = [w_vals[key][1] for key in w_vals]
        scale = calculate_print_scale(widths)
        sheet.page_setup.scale = scale
        return True

    def update_fits(self, sheet, fit_width=1, fit_height=1):
        """ Fits page width and page(s) heights.
        """
        sheet.page_setup.pageSetupType = 'fitToPage'
        sheet.page_setup.scale = None
        sheet.page_setup.fitToWidth = fit_width		# fit to 1 page wide
        sheet.page_setup.fitToHeight = fit_height	# fit to 1 pages tall
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        return True

    def update_new_scale(self, sheet, new_scale=100, paper_size=None):
        """ Fits page with a new (manual) scale value.
        """
        sheet.page_setup.scale = new_scale
        sheet.page_setup.fitToWidth = None
        sheet.page_setup.fitToHeight = None
        if paper_size is None:
            return False
        sheet.page_setup.paperSize = paper_size
        return True

    def update_headers(self, new_sheet, hdr=None, ftr=None):
        h_obj = HeaderFooter()
        left, center, right = (None, None, None) if hdr is None else hdr
        left_f, center_f, right_f = (None, None, None) if ftr is None else ftr
        if left:
            h_obj.left_header = left
        if center:
            h_obj.center_header = center
        if center:
            h_obj.right_header = right
        if left_f:
            h_obj.left_footer = left_f
        if center_f:
            h_obj.center_footer = center_f
        if center_f:
            h_obj.right_footer = right_f
        # Force Excel to recognize the header/footer
        h_obj.differentFirst = False
        h_obj.differentOddEven = False
        new_sheet.header_footer = h_obj
        # Apply to oddHeader to ensure Excel picks it up
        new_sheet.oddHeader.left.text = self._replaced(left, new_sheet) or ""
        new_sheet.oddHeader.center.text = self._replaced(center, new_sheet) or ""
        new_sheet.oddHeader.right.text = self._replaced(right, new_sheet) or ""
        # The same for footer
        new_sheet.oddFooter.left.text = self._replaced(left_f, new_sheet) or ""
        new_sheet.oddFooter.center.text = self._replaced(center_f, new_sheet) or ""
        new_sheet.oddFooter.right.text = self._replaced(right_f, new_sheet) or ""
        # Returns True if any of header elements is not None
        return any(hdr)

    def _replaced(self, astr, sheet):
        title = sheet.title
        astr = astr.replace("$101", title)
        return astr


def copycat(sheet_source, sheet_new=None, bare_width=25):
    """ Copy a sheet into a new Workbook() sheet. """
    sheet_new.title = sheet_source.title
    larg = {}
    dct = {
        "widths": larg,
    }
    # Copy content and formatting
    for row in sheet_source.iter_rows():
        for cell in row:
            if not isinstance(cell, Cell):
                continue
            new_cell = sheet_new.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                update_cell_style(new_cell, cell)
    last = bare_width	# any reasonable width, if none before was found
    # Copy column widths
    for col_idx in range(1, sheet_source.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in sheet_source.column_dimensions:
            width = sheet_source.column_dimensions[col_letter].width
            ori = width
        else:
            ori = None
            width = last
        larg[col_letter] = (
            ori,
            width,
            conv_x_to_pixels(width),
        )
        last = width
    for col_letter in sorted(larg):
        largo = larg[col_letter][1]
        sheet_new.column_dimensions[col_letter].width = largo
    # Copy row heights
    for row_idx, row_dim in sheet_source.row_dimensions.items():
        sheet_new.row_dimensions[row_idx].height = row_dim.height
    copy_page_layout(sheet_source, sheet_new)
    return dct


def copy_page_layout(source_sheet, target_sheet):
    """ Copy page setup and margins from source_sheet to target_sheet.
    Includes orientation, paper size, scaling, and margins.
    """
    # Page setup
    target_sheet.page_setup.orientation = source_sheet.page_setup.orientation
    target_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
    target_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth
    target_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
    target_sheet.page_setup.scale = source_sheet.page_setup.scale
    # Page margins
    target_sheet.page_margins.left = source_sheet.page_margins.left
    target_sheet.page_margins.right = source_sheet.page_margins.right
    target_sheet.page_margins.top = source_sheet.page_margins.top
    target_sheet.page_margins.bottom = source_sheet.page_margins.bottom
    target_sheet.page_margins.header = source_sheet.page_margins.header
    target_sheet.page_margins.footer = source_sheet.page_margins.footer


def update_cell_style(new_cell, cell):
    new_cell.font = copy.copy(cell.font)
    new_cell.border = copy.copy(cell.border)
    new_cell.fill = copy.copy(cell.fill)
    new_cell.number_format = copy.copy(cell.number_format)
    new_cell.protection = copy.copy(cell.protection)
    new_cell.alignment = copy.copy(cell.alignment)
    return True


def conv_x_to_pixels(x_width, factor=7) -> int:
    """ Converts Excel widths stored into pixels, approximately. """
    pix = round(x_width * factor + 5)
    return pix


def calculate_print_scale(column_widths, max_pixels=756):
    """ Given a list of Excel column widths, calculate the print scale percentage
    to fit within max_pixels (default: A4 landscape printable width).
    """
    total_pixels = sum(conv_x_to_pixels(width) for width in column_widths)
    scale = min(100, round((max_pixels / total_pixels) * 100))
    return scale


# Main script
if __name__ == "__main__":
    print("Please import me!")
    tester("test-input.xlsx")
